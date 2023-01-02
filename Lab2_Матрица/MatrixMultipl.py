import threading

# Импортируем модуль измерения времени
import time
# Импортируем модуль для создания многопоточности
from multiprocessing import Pool
# для записи в файл Excel
from openpyxl import load_workbook

import pandas as pd
from openpyxl.reader.excel import load_workbook

# Чтение матрицы из базы данных файла csv
matrixA = pd.read_csv('MatrixA.csv', header=None)
# Вывести матрицу A в консоль
print("")
print("Матрица A")
print(matrixA)
print("")

# Создание матрицы B. Для удобства принимаем матриу B равной матрице A.
matrixB = matrixA
print("Матрица B")
print(matrixB)
print("")


# Умножение матриц
# Функция для перемножения матриц
def matrix_div(matrixA, matrixB):
    return matrixA.dot(matrixA)


# цикл для запуска умножения матриц необходимое количество раз
for number in range(10):
    # задаем начальное время
    t0 = time.time()

    # t1 = threading.Thread(target=matrix_div, args=(matrixA, matrixB))
    # t2 = threading.Thread(target=matrix_div, args=(matrixA, matrixB))
    # t3 = threading.Thread(target=matrix_div, args=(matrixA, matrixB))
    # t4 = threading.Thread(target=matrix_div, args=(matrixA, matrixB))
    # t5 = threading.Thread(target=matrix_div, args=(matrixA, matrixB))
    # t6 = threading.Thread(target=matrix_div, args=(matrixA, matrixB))
    #
    # t1.start()
    # t1.join()
    # t2.start()
    # t2.join()
    # t3.start()
    # t3.join()
    # t4.start()
    # t4.join()
    # t5.start()
    # t5.join()
    # t6.start()
    # t6.join()

    # вызов ранее заданной функции для умножения матриц
    matrixC = matrix_div(matrixA, matrixB)
    print('Матрица C = Матрица A x Матрица B:')
    print(matrixC)

    print("")
    print("Количество активных потоков - ", threading.active_count())

    # Расчет времени
    print("")
    t1 = time.time() - t0
    print("Затраченное на умножение матриц время:", round(t1 * 10 ** 3, 0), "мс, или", round(t1, 3), "с")
    print("")

    # Попытка перемножения матриц "физически", не через библиотеку
    # X = matrixA
    # Y = matrixA
    # matrixD = [[sum(a * b for a, b in zip(X_row, Y_col)) for Y_c ol in zip(*Y)] for X_row in X]

    # print("Матрица D = Матрица A x Матрица B:")
    # for x in matrixD:
    #     print(x)

    # Запись времени в файл Excel
    fnTime = 'MatrixTime.xlsx'
    wb = load_workbook(fnTime)
    ws = wb["Лист1"] # не применяем многопоточность
#   ws = wb["1поток"]
#   ws = wb["2потока"]
#   ws = wb["4потока"]
#   ws = wb["6потоков"]
    # добавление по порядку ПОСЛЕ уже записанного
    ws.append([round(t1, 5)])

    wb.save(fnTime)
    wb.close
